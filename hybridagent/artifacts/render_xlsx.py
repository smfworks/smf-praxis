"""Optional XLSX renderer for professional artifact workbooks."""
from __future__ import annotations

import io
from datetime import datetime

from hybridagent.artifacts.models import ArtifactDocument, FigureBlock, TableBlock
from hybridagent.artifacts.render_common import (
    ArtifactRenderError,
    checked_assets,
    image_kind,
    normalize_zip_package,
    require_backend,
)

_FIXED_TIME = datetime(2000, 1, 1)


def _safe_cell(value: object) -> object:
    if type(value) is str and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_sheet(value: str, used: set[str]) -> str:
    base = "".join("_" if char in "[]:*?/\\" else char for char in value).strip() or "Table"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f"_{index}"
        candidate = base[: 31 - len(suffix)] + suffix
        index += 1
    used.add(candidate.casefold())
    return candidate


def render_xlsx(
    document: ArtifactDocument, assets: dict[str, bytes] | None = None
) -> bytes:
    require_backend("openpyxl", "openpyxl")
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.drawing.image import Image  # type: ignore[import-untyped]
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]

    resolved = checked_assets(document, assets)
    workbook = Workbook()
    workbook.properties.title = document.metadata.title
    workbook.properties.subject = document.metadata.subject
    workbook.properties.creator = document.metadata.created_by
    workbook.properties.lastModifiedBy = "Praxis Artifact Studio"
    workbook.properties.created = _FIXED_TIME
    workbook.properties.modified = _FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    used: set[str] = set()
    overview = workbook.active
    overview.title = _safe_sheet("Overview", used)
    overview.append(["Field", "Value"])
    for key, value in (
        ("Artifact ID", document.artifact_id),
        ("Title", document.metadata.title),
        ("Subject", document.metadata.subject),
        ("Document type", document.metadata.document_type),
        ("Language", document.metadata.language),
        ("Classification", document.metadata.confidentiality),
        ("Organization", document.metadata.organization_id),
        ("Workspace", document.metadata.workspace_id),
        ("Canonical SHA-256", document.content_hash()),
    ):
        overview.append([key, _safe_cell(value)])
    overview.freeze_panes = "A2"
    overview.print_title_rows = "1:1"

    sections = workbook.create_sheet(_safe_sheet("Sections", used))
    sections.append(["Section", "Level", "Block", "Type", "Content", "Citations"])
    table_number = 0
    figure_number = 0
    for section in (*document.sections, *document.appendices):
        for block in section.blocks:
            content = getattr(block, "text", "")
            if hasattr(block, "items"):
                content = "\n".join(block.items)
            sections.append([_safe_cell(value) for value in (
                section.title, section.level, block.block_id,
                type(block).__name__, content,
                ", ".join(getattr(block, "citation_ids", ())),
            )])
            if type(block) is TableBlock:
                table_number += 1
                sheet = workbook.create_sheet(
                    _safe_sheet(block.caption or f"Table {table_number}", used)
                )
                sheet.append([_safe_cell(value) for value in block.columns])
                for row in block.rows:
                    sheet.append([_safe_cell(value) for value in row])
                sheet.freeze_panes = "A2"
                sheet.print_title_rows = "1:1"
            elif type(block) is FigureBlock:
                figure_number += 1
                sheet = workbook.create_sheet(_safe_sheet(f"Figure {figure_number}", used))
                sheet.append(["Asset ID", _safe_cell(block.asset_id)])
                sheet.append(["Alternative text", _safe_cell(block.alt_text)])
                sheet.append(["Caption", _safe_cell(block.caption)])
                payload = resolved[block.asset_id]
                if image_kind(payload) == "svg":
                    raise ArtifactRenderError("XLSX rendering requires PNG or JPEG figure assets")
                image = Image(io.BytesIO(payload))
                image.width = min(float(image.width), 720.0)
                image.height = min(float(image.height), 480.0)
                sheet.add_image(image, "A5")
    sections.freeze_panes = "A2"
    sections.auto_filter.ref = sections.dimensions

    sources = workbook.create_sheet(_safe_sheet("Sources", used))
    sources.append(["Source", "Version", "SHA-256", "Title", "URI"])
    for source in document.sources:
        sources.append([_safe_cell(value) for value in (
            source.source_id, source.source_version_id, source.content_hash,
            source.title, source.canonical_uri,
        )])
    citations = workbook.create_sheet(_safe_sheet("Citations", used))
    citations.append(["Citation", "Source", "Spans", "Claims", "Pinpoint"])
    for citation in document.citations:
        citations.append([_safe_cell(value) for value in (
            citation.citation_id, citation.source_id, ", ".join(citation.span_ids),
            ", ".join(citation.claim_ids), citation.pinpoint,
        )])
    governance = workbook.create_sheet(_safe_sheet("Governance", used))
    governance.append(["Kind", "ID", "Actor", "State/meaning", "Timestamp"])
    for revision in document.revisions:
        governance.append([_safe_cell(value) for value in
                           ("revision", revision.revision_id, revision.author_id,
                            revision.summary, revision.created_at)])
    for review in document.reviews:
        governance.append([_safe_cell(value) for value in
                           ("review", review.review_id, review.reviewer_id,
                            review.decision, review.reviewed_at)])
    for signature in document.signatures:
        governance.append([_safe_cell(value) for value in
                           ("signature", signature.signature_id, signature.signer_id,
                            signature.meaning, signature.signed_at)])

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = sheet.freeze_panes or "A2"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    return normalize_zip_package(output.getvalue())
